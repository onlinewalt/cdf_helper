"""End-to-end tests of the Flask web app via the test client.

Hermetic: the template is read from the committed repo fixture and the
source workbook is built in memory and uploaded through the client, so these
tests need no root data files and are stable in CI (the repo's sample data
files are gitignored).
"""
import io
import re

import openpyxl
import pytest

import webapp


def _template_bytes():
    """Bytes of the committed template workbook (glob is locale-safe)."""
    from pathlib import Path

    for f in Path(".").glob("*报关清单*.xlsx"):
        return io.BytesIO(f.read_bytes())
    pytest.skip("no committed template workbook found in repo root")


def _make_source_bytes():
    """A minimal packing-list source workbook built in memory."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Item      Quantity(Unit)         Particulars"
    ws["A2"] = "1"
    ws["B2"] = "1  PCE"
    ws["C2"] = "消防泵出海阀(蝶阕)"
    ws["B3"] = "Type:5k250"
    ws["A4"] = "2"
    ws["B4"] = "2 PCE"
    ws["C4"] = "液压法兰式蝶阙"
    ws["B6"] = "** End of Listing **"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _generate(client, tpl, src, *, use_ai=False, api_key="", save_key=False, follow=False):
    data = {
        "vessel": "测试船",
        "port": "测试港",
        "date": "2026-08-20",
        "include_spec": "on",
        "template_upload": (io.BytesIO(tpl.getvalue()), "template.xlsx"),
        "sources_upload": [(io.BytesIO(src.getvalue()), "source.xlsx")],
    }
    if use_ai:
        data["use_ai"] = "on"
        data["api_key"] = api_key
        if save_key:
            data["save_key"] = "on"
    return client.post(
        "/generate",
        data=data,
        content_type="multipart/form-data",
        follow_redirects=follow,
    )


def test_index_renders(client):
    r = client.get("/")
    assert r.status_code == 200
    assert "报关清单生成器" in r.get_data(as_text=True)


def test_generate_and_download(client):
    tpl, src = _template_bytes(), _make_source_bytes()
    r = _generate(client, tpl, src)
    assert r.status_code == 200
    body = r.get_data(as_text=True)
    m = re.search(r'href="(/download/[^"]+)"', body)
    assert m, "no download link found"
    dl = client.get(m.group(1).replace("&amp;", "&"))
    assert dl.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(dl.data))
    ws = wb["Sheet1"]
    assert "船名：" in str(ws["A1"].value)
    assert ws.max_row >= 4  # title + header + >=1 item + total row


def test_generate_error_when_no_sources(client):
    tpl = _template_bytes()
    r = client.post(
        "/generate",
        data={
            "template_upload": (io.BytesIO(tpl.getvalue()), "template.xlsx"),
            "vessel": "测试船",
            "port": "港",
            "date": "2026-08-20",
        },
        content_type="multipart/form-data",
        follow_redirects=True,
    )
    assert "请至少选择一个" in r.get_data(as_text=True)


def test_generate_with_ai(monkeypatch, client):
    def fake_enrich(parts, api_key, on_status=None):
        for p in parts:
            if p.weight is None:
                p.weight = 9.9
            if p.price is None:
                p.price = 88.8
        if on_status:
            on_status("（模拟）DeepSeek 估算 2 条")
        return {"requested": 2, "filled": 2, "from_cache": 0, "errors": 0}

    monkeypatch.setattr(webapp, "enrich_parts", fake_enrich)
    tpl, src = _template_bytes(), _make_source_bytes()
    r = _generate(client, tpl, src, use_ai=True, api_key="sk-test")
    assert r.status_code == 200
    assert "DeepSeek" in r.get_data(as_text=True)


def test_generate_ai_without_key(client):
    tpl, src = _template_bytes(), _make_source_bytes()
    r = _generate(client, tpl, src, use_ai=True, api_key="", follow=True)
    assert "API Key" in r.get_data(as_text=True)


if __name__ == "__main__":
    import sys
    sys.exit(pytest.main([__file__, "-v"]))
