"""Parse spare-parts source Excel files into a normalized list of parts.

A Part is a dataclass with:
    name      - part name (名称 / 物资名称 / Particulars)
    qty       - quantity (数量)
    unit      - unit (单位)
    type      - model / spec (规格 / Type / 物资规格/品牌)
    weight    - weight in KG (重量), optional
    price     - unit price in RMB (单价), optional

Supported source layouts:
1. 中文签收单/物料清单：表头含 名称 与 数量（自动识别）。
2. 英文 Receipt/Packing List（如远通海事 Yuantong）：表头含 Item/Quantity(Unit)/Particulars，
   数量形如 "N PCE"，以 "** End of Listing **" 结尾；多 sheet 各自为一张装箱单。
Both .xlsx (openpyxl) and legacy .xls (xlrd) files are supported.
"""

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

try:
    import xlrd
except ImportError:
    xlrd = None


@dataclass
class Part:
    name: str
    qty: float = 1.0
    unit: str = "个"
    type: Optional[str] = None
    weight: Optional[float] = None
    price: Optional[float] = None


HEADER_ALIASES = {
    "name": ("备件名称", "名称", "备件", "品名", "物料名称", "物资名称", "名称及规格", "名称/规格", "part name", "name", "description", "description of goods", "des", "particulars", "particular"),
    "type": ("型号", "规格", "型号规格", "规格型号", "物资规格", "物资规格/品牌", "规格/品牌", "type", "spec", "model", "品牌", "part no", "part no.", "parts no", "parts no."),
    "qty": ("数量", "件数", "qty", "quantity", "数量/套", "数量  ", "q'ty", "qtt"),
    "unit": ("单位", "unit", "计量单位", "单位  "),
    "weight": ("重量", "重量(kg)", "weight", "单重", "毛重"),
    "price": ("单价", "单价/rmb", "unit price", "price", "价格"),
}

_CONTAINS_RULES = (
    # (needle, field) checked in order; a header cell that contains any needle maps to the field
    ("名称及规格", "name"),
    ("名称", "name"),
    ("品名", "name"),
    ("description", "name"),
    ("des", "name"),
    ("规格", "type"),
    ("型号", "type"),
    ("品牌", "type"),
    ("part no", "type"),
    ("数量", "qty"),
    ("件数", "qty"),
    ("q'ty", "qty"),
    ("qty", "qty"),
    ("单位", "unit"),
    ("重量", "weight"),
    ("单价", "price"),
    ("价格", "price"),
)

# ---- packing-list (English Receipt/Packing List) patterns -------------
_QTY_RE = re.compile(
    r"(?<!\d)(\d+(?:\.\d+)?)\s*"
    r"(PCE|PCS|PC|SETS|SET|SHEET|PKT|MTR|PAIR|CASE|CAN|EA|BAG|ROLL|BIL)(?!\w)",
    re.IGNORECASE,
)
_END_RE = re.compile(r"end\s+of\s+listing", re.IGNORECASE)

# Detects a line starting with a sequence number or checkbox+number (multi-part format)
_SEQ_PREFIX_RE = re.compile(r"^[✓✗☐☑\u2713\u2717\u2714\u2716\u2718\u2610\u2611\u2612工\s]*\d+(?:\.\d+)?\s")

_IRRELEVANT_RE = re.compile(
    r"^\s*(?:"
    r"\*\*|\(\d+\)|（\d+）|Equipment|Item|Quantity|Particulars|Part\s*No|Serial\s*No|"
    r"Cus\s*Ref|Status|Gross\s*Weight|Order\s*No|Order\s*Date|Msg\s*No|Vessel|Owner\s*Ref|"
    r"Department|Delivery|Package|Coordinator|Email|Tel|Date|Signature|Issued|Page|"
    r"RI#|Shipment#|http"
    r")",
    re.IGNORECASE,
)

_UNIT_WORDS = frozenset(
    w.upper()
    for w in (
        "PCE", "PCS", "PC", "SET", "SETS", "SHEET", "PKT", "MTR", "PAIR",
        "CASE", "CAN", "EA", "BAG", "ROLL", "个", "件", "套", "台", "支",
        "根", "张", "块", "盒", "米", "只", "片",
    )
)

_FOOTER_KEYWORDS = (
    "signature",
    "盖章", "蓋章", "蓝章", "藍章",
    "签收", "簽收", "簸收",
    "issued by",
    "date 日期",
    "htt",
)


def _trim_footer(text: str) -> str:
    """Cut trailing footer fragments (signatures, URLs) glued onto a name."""
    low = text.lower()
    idx = len(text)
    for kw in _FOOTER_KEYWORDS:
        i = low.find(kw)
        if i != -1:
            idx = min(idx, i)
    if idx < len(text):
        text = text[:idx].strip(" -–—")
    return text


def _norm(s: str) -> str:
    return str(s).strip().replace(" ", "").replace("\u3000", "").lower()


def _clean(value) -> str:
    """Collapse whitespace/newlines of a cell value into a single line."""
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    return " ".join(text.split())


def _clean_name(text: str) -> str:
    """Drop pure-number / stray-unit / separator tokens from a name fragment."""
    tokens = []
    for t in text.split():
        t = t.strip(":：")
        if re.fullmatch(r"[\d.,\-]+", t):
            continue
        if re.fullmatch(r"[=\-*]+", t):
            continue
        if t.upper() in _UNIT_WORDS:
            continue
        tokens.append(t)
    return " ".join(tokens).strip(" -–—")


def _split_type(text: str):
    """Return (text_without_type, type_part). Splits a "Type:xxx" tail out of a cell."""
    m = re.search(r"Type\s*[:：]", text, re.IGNORECASE)
    if not m:
        return text, None
    before = text[: m.start()].strip(" -–—")
    after = text[m.end():].strip()
    return before, after or None


def _is_irrelevant(text: str) -> bool:
    if text.startswith(":") or text.endswith(":"):
        return True
    if re.match(r"dwg\.?", text, re.IGNORECASE):
        return True
    if re.fullmatch(r"[\d.,\-()（）\s]+", text):
        return True
    return bool(_IRRELEVANT_RE.match(text))


_CHECKBOX_CHARS = "✓✗☐☑\u2713\u2717\u2714\u2716\u2718\u2610\u2611\u2612工"


def _strip_seq(text: str) -> str:
    """Strip leading sequence number / checkbox char and trailing sequence number
    from padded part names (e.g. Sheet10 '工  INTERMEDIATE RELAY   1' → 'INTERMEDIATE RELAY')."""
    text = text.strip()
    # Strip leading checkbox character followed by whitespace
    m = re.match(f'^[{_CHECKBOX_CHARS}\\s]+', text)
    text = text[m.end():] if m else text
    # Strip leading number + 2+ spaces (sequence number prefix)
    m = re.match(r"^\d+(?:\.\d+)?[\s\u3000]{2,}", text)
    text = text[m.end():] if m else text
    # Strip trailing spaces + number (sequence number suffix)
    m = re.search(r"[\s\u3000]{2,}\d+(?:\.\d+)?\s*$", text)
    text = text[: m.start()] if m else text
    return text.strip()


class _Cell:
    """Minimal unified cell wrapper over xlrd / openpyxl cells."""

    __slots__ = ("value", "column", "row")

    def __init__(self, value, column, row):
        self.value = value
        self.column = column
        self.row = row


class _Sheet:
    """Minimal unified sheet wrapper supporting iter_rows() and cell()."""

    def __init__(self, source):
        self._source = source
        if source.__class__.__module__.startswith("xlrd"):
            self._kind = "xls"
        else:
            self._kind = "xlsx"

    @property
    def name(self) -> str:
        if self._kind == "xls":
            return str(self._source.name)
        return str(self._source.title)

    @property
    def last_row(self) -> int:
        if self._kind == "xls":
            return int(self._source.nrows)
        return int(self._source.max_row)

    def iter_rows(self):
        if self._kind == "xls":
            for r in range(self._source.nrows):
                yield [
                    _Cell(self._source.cell_value(r, c), c + 1, r + 1)
                    for c in range(self._source.ncols)
                ]
        else:
            for row in self._source.iter_rows():
                yield [_Cell(c.value, c.column, c.row) for c in row]

    def cell(self, row, column):
        if self._kind == "xls":
            return _Cell(self._source.cell_value(row - 1, column - 1), column, row)
        return self._source.cell(row, column)


def _open_sheets(path: Path):
    path = Path(path)
    if path.suffix.lower() == ".xls":
        if xlrd is None:
            raise RuntimeError("需要 xlrd 库来读取 .xls 文件：pip install xlrd")
        wb = xlrd.open_workbook(str(path))
        return [_Sheet(s) for s in wb.sheets()]
    wb = load_workbook(path, data_only=True)
    return [_Sheet(ws) for ws in wb.worksheets]


# ---- standard (Chinese header) parsing --------------------------------

def _find_header_row(sheet):
    """Return (header_row_idx, {field: col_idx}) for the first row that looks like a parts header."""
    for cells in sheet.iter_rows():
        mapping = {}
        for cell in cells:
            if cell.value is None:
                continue
            text = _norm(cell.value)
            if not text:
                continue
            matched = False
            for field, aliases in HEADER_ALIASES.items():
                for alias in aliases:
                    if _norm(alias) == text:
                        if field not in mapping:
                            mapping[field] = cell.column
                        matched = True
                        break
                if matched:
                    break
            if matched:
                continue  # exact alias wins; do not also contains-match this cell
            for needle, field in _CONTAINS_RULES:
                if field in mapping:
                    continue
                if needle in text:
                    mapping[field] = cell.column
                    # Do NOT break — allow multiple field matches in one cell
                    # (needed for packed header format where all headers are in one cell)
        if "name" in mapping and "qty" in mapping:
            return cells[0].row, mapping
    return None, None


_SIGNATURE_KEYWORDS = (
    "签字", "签收", "盖章", "供船", "签收日期", "交货地址", "请核对", "核对",
    "日 期", "signed", "signature", "received", "place of supply",
    "date of supply", "please check",
    "目的地", "供货单位", "收货人", "经办人", "邮箱", "日期", "单号",
)


def _is_signature(text: str) -> bool:
    low = " ".join(str(text).lower().split())
    return any(k in low for k in _SIGNATURE_KEYWORDS)


def _split_lines(val) -> list:
    """Split a cell value by newlines, returning a list of stripped strings."""
    if val is None:
        return []
    return [s.strip() for s in str(val).split("\n")]


def _find_unit_column(cells, name_col, qty_col, unit_col, type_col):
    """When unit_col is missing or equals qty_col, search data cells for a
    column holding multi-line non-numeric values (units like PC, SET, MTR)."""
    if unit_col and unit_col != qty_col:
        return unit_col
    for cell in cells:
        if cell.column in (name_col, qty_col, type_col):
            continue
        val = _to_text(cell.value)
        if not val or "\n" not in val:
            continue
        lines = [s.strip() for s in val.split("\n")]
        lines = [l for l in lines if l]
        if lines and all(not _to_number(l) and len(l) <= 12 for l in lines):
            return cell.column
    return unit_col


def _resolve_cols(mapping):
    """Resolve column numbers from a header *mapping* dict.

    Returns ``(name_col, qty_col, type_col, unit_col, weight_col, price_col)``.
    *type_col* is forced to ``None`` when it overlaps *name_col*, since a single
    column cannot hold both the part name and its spec/model.
    """
    name_col = mapping["name"]
    qty_col = mapping["qty"]
    type_col = mapping.get("type")
    if type_col == name_col:
        type_col = None
    return (
        name_col,
        qty_col,
        type_col,
        mapping.get("unit"),
        mapping.get("weight"),
        mapping.get("price"),
    )


def _cell_value(cells, col):
    """Return the value of the cell at 1-based column *col*, or ``None``.

    Returns ``None`` when *col* is falsy (missing column), so callers can write
    ``_cell_value(cells, col)`` instead of the verbose
    ``cells[col - 1].value if col else None``.
    """
    if not col:
        return None
    return cells[col - 1].value


def _expand_lines(lines, count):
    """Broadcast a single-element *lines* list to *count* entries.

    A scalar column (one value) that applies to every part in a multi-line cell
    is replicated so each part gets its own value; lists that already match or
    exceed *count* are returned unchanged.
    """
    if len(lines) == 1 and count > 1:
        return lines * count
    return lines


def _split_and_expand(cells, col, count):
    """Retrieve, split (by newline) and expand a cell's lines to *count* entries."""
    return _expand_lines(_split_lines(_cell_value(cells, col)), count)


def _parse_multiline_row(cells, mapping, path, warn) -> list:
    """Parse a row where the name column has multiple newline-separated parts
    (each line prefixed with a sequence number)."""
    name_col, qty_col, type_col, unit_col, weight_col, price_col = _resolve_cols(mapping)

    name_lines = [l for l in _split_lines(_cell_value(cells, name_col)) if l]
    if not name_lines:
        return []

    n = len(name_lines)
    qty_lines = _split_and_expand(cells, qty_col, n)
    actual_unit_col = _find_unit_column(cells, name_col, qty_col, unit_col, type_col)
    unit_lines = _split_and_expand(cells, actual_unit_col, n)
    type_lines = _split_and_expand(cells, type_col, n)

    # weight & price are constant across all lines in this row — compute once
    row_num = cells[0].row
    weight = _to_number(_cell_value(cells, weight_col)) if weight_col else None
    price = _to_number(_cell_value(cells, price_col)) if price_col else None

    parts = []
    for idx, line in enumerate(name_lines):
        name = _clean(_strip_seq(line))
        if not name or _is_signature(name):
            continue
        qty = _to_number(qty_lines[idx]) if idx < len(qty_lines) else None
        if qty is None:
            if warn:
                warn(f"{path.name} 第 {row_num} 行第 {idx+1} 个备件缺少数量，按 1 处理: {name}")
            qty = 1.0
        unit = unit_lines[idx].strip() if idx < len(unit_lines) else None
        part_type = type_lines[idx].strip() if idx < len(type_lines) else None
        parts.append(
            Part(
                name=name,
                qty=qty,
                unit=unit or "个",
                type=part_type,
                weight=weight,
                price=price,
            )
        )
    return parts


def _parse_standard(sheet, mapping, header_row, path, warn) -> list:
    name_col, qty_col, type_col, unit_col, weight_col, price_col = _resolve_cols(mapping)

    def make_part(cells, name, qty, unit):
        return Part(
            name=name,
            qty=qty,
            unit=unit or "个",
            type=_clean(_cell_value(cells, type_col)) if type_col else None,
            weight=_to_number(_cell_value(cells, weight_col)) if weight_col else None,
            price=_to_number(_cell_value(cells, price_col)) if price_col else None,
        )

    parts = []
    for cells in sheet.iter_rows():
        if cells[0].row <= header_row:
            continue
        raw_name = _cell_value(cells, name_col)
        if raw_name is None:
            continue

        # Multi-line cell: check if it's a multi-part format (lines start with seq numbers)
        if "\n" in str(raw_name):
            raw_lines = str(raw_name).split("\n")
            seq_count = sum(1 for l in raw_lines if _SEQ_PREFIX_RE.match(l.strip()))
            if seq_count >= 2:
                parts.extend(_parse_multiline_row(cells, mapping, path, warn))
                continue
            # Single part with multi-line name — fall through; _clean collapses newlines
        name = _clean(_strip_seq(str(raw_name)))

        if not name:
            continue
        if _is_signature(name):
            continue  # signature / footer rows must not become parts

        qty = _to_number(_cell_value(cells, qty_col))
        if qty is None:
            if warn:
                warn(f"{path.name} 第 {cells[0].row} 行缺少数量，按 1 处理: {name}")
            qty = 1.0
        unit = _to_text(_cell_value(cells, unit_col)) if unit_col else None
        parts.append(make_part(cells, name, qty, unit))
    return parts


# ---- packed single-column format parsing --------------------------------

_PACKED_QTY_RE = re.compile(r'^[\d.]+$')
_CODE_PATTERN = re.compile(r'[\u00d7/]|dwg|code|drg', re.IGNORECASE)


def _is_packed_format(mapping) -> bool:
    """True when all mapped fields point to the same column (space-separated packed format)."""
    cols = set(v for v in mapping.values() if v is not None)
    return len(cols) == 1


def _collect_packed_lines(sheet, header_row, packed_col):
    """Collect data lines from the header cell (may contain header + data) and subsequent rows."""
    lines = []

    # 1) Check if the header cell itself contains data lines after the header line
    header_cell = sheet.cell(header_row, packed_col)
    if header_cell.value is not None:
        cell_lines = _split_lines(header_cell.value)
        header_line_idx = None
        for i, line in enumerate(cell_lines):
            norm = _norm(line)
            if any(n in norm for n in ("序号", "名称", "name", "品名", "qty", "quantity")):
                header_line_idx = i
                break
        if header_line_idx is not None:
            for line in cell_lines[header_line_idx + 1:]:
                line = line.strip()
                if line and not _is_signature(line):
                    lines.append(line)

    # 2) Process subsequent rows
    for row_idx in range(header_row + 1, sheet.last_row + 1):
        cell = sheet.cell(row_idx, packed_col)
        if cell.value is None:
            continue
        for line in _split_lines(cell.value):
            line = line.strip()
            if line and not _is_signature(line):
                lines.append(line)

    return lines


def _parse_packed_line(line, path, warn):
    """Parse a single packed data line into a Part.

    Format: 序号  设备  名称  编号  单位  数量  备注
    Values are separated by 2+ spaces. Some fields may be missing.
    """
    values = re.split(r'\s{2,}', line)
    values = [v.strip() for v in values if v.strip()]

    if not values:
        return None

    # Remove seq number (first value if it's a simple number)
    if values and _PACKED_QTY_RE.match(values[0]):
        values = values[1:]

    if not values:
        return None

    # Extract qty (last value if it's a simple number)
    qty = 1.0
    if values and _PACKED_QTY_RE.match(values[-1]):
        qty = float(values[-1])
        values = values[:-1]

    if not values:
        if warn:
            warn(f"{path.name}: 无法从行解析名称: {line!r}")
        return None

    # Identify code/type values
    type_parts = []
    remaining = []
    for v in values:
        if _CODE_PATTERN.search(v):
            type_parts.append(v)
        else:
            remaining.append(v)

    type_str = ' '.join(type_parts) if type_parts else None

    # Remaining values: first is equipment (if Chinese), rest is name
    name_parts = []
    for v in remaining:
        name_parts.append(v)

    name = ' '.join(name_parts)
    if not name:
        return None

    return Part(
        name=_clean(_strip_seq(name)),
        qty=qty,
        unit='个',
        type=_clean(type_str) if type_str else None,
    )


def _parse_packed_sheet(sheet, mapping, header_row, path, warn) -> list:
    """Parse sheets where all header fields are in a single cell (packed format).

    Data rows have fields separated by 2+ spaces within a single cell.
    Handles both multi-line cell format (header and data in same cell)
    and multi-row format (header and data in separate rows).
    """
    cols = [v for v in mapping.values() if v is not None]
    if not cols:
        return []
    packed_col = cols[0]

    data_lines = _collect_packed_lines(sheet, header_row, packed_col)

    parts = []
    for line in data_lines:
        part = _parse_packed_line(line, path, warn)
        if part:
            parts.append(part)

    return parts


# ---- packing-list (English Receipt/Packing List) parsing --------------

def _find_packing_header(sheet):
    """Return (header_row, exclude_cols).

    A packing header is a row containing both "item" and "quantity" (the
    particulars/description label may be missing or named "Description").
    Columns labelled Part No / Serial No / Dwg.* are metadata and excluded
    from name extraction.
    """
    for cells in sheet.iter_rows():
        texts = [_clean(c.value) for c in cells if c.value is not None]
        joined = " ".join(texts).lower()
        qty_found = "quantity" in joined or "qty" in joined or "q'ty" in joined or "qtt" in joined
        if "item" not in joined or not qty_found:
            continue
        exclude = set()
        for c in cells:
            t = _clean(c.value).lower()
            if not t:
                continue
            if any(k in t for k in ("item", "quantity", "particulars", "description")):
                continue  # merged main header cell; not a metadata column
            if "part no" in t or "serial no" in t or "dwg" in t:
                exclude.add(c.column)
        return cells[0].row, exclude
    return None, set()


def _row_parts(cells, exclude_cols=()):
    """Extract (qty, unit, name_parts, type_parts) from one row of cells."""
    qty, unit = None, None
    name_parts, type_parts = [], []
    for cell in cells:
        if cell.column in exclude_cols:
            continue
        text = _clean(cell.value)
        if not text:
            continue
        base, type_part = _split_type(text)
        if type_part:
            type_parts.append(type_part)
        if not base:
            continue
        m = _QTY_RE.search(base)
        if m:
            if qty is None:
                qty = float(m.group(1))
                unit = m.group(2)
            rest = base[: m.start()] + base[m.end():]
            rest = _clean_name(rest)
            if rest and not _is_irrelevant(rest):
                name_parts.append(rest)
            continue
        if _is_irrelevant(base):
            continue
        name_parts.append(_clean_name(base))
    return qty, unit, name_parts, type_parts


def _parse_packing_sheet(sheet, path, warn) -> list:
    header_row, exclude_cols = _find_packing_header(sheet)
    if header_row is None:
        if warn:
            warn(f"{path.name}/{sheet.name}: 未找到 Item/Quantity 表头，跳过该表")
        return []

    rows = list(sheet.iter_rows())
    end = sheet.last_row + 1
    for idx in range(header_row, sheet.last_row + 1):
        text = " ".join(_clean(c.value) for c in rows[idx - 1] if c.value is not None)
        if _END_RE.search(text):
            # "End of Listing" is sometimes embedded inside a data row's cell
            # (e.g. a Serial No. cell carrying the marker between sections)
            # rather than sitting on its own footer row. A genuine footer row
            # carries no item quantity; only truncate on those.
            if _QTY_RE.search(text):
                continue
            end = idx
            break

    items = []
    i = header_row
    while i < end:
        cells = rows[i - 1]
        qty, unit, name_parts, type_parts = _row_parts(cells, exclude_cols)
        if qty is None:
            i += 1
            continue

        # look ahead: continuation / Type lines until the next qty row or end
        j = i + 1
        while j < end:
            nqty, _, nname, ntype = _row_parts(rows[j - 1], exclude_cols)
            if nqty is not None:
                break
            text = " ".join(_clean(c.value) for c in rows[j - 1] if c.value is not None)
            if _END_RE.search(text):
                break
            name_parts.extend(nname)
            type_parts.extend(ntype)
            j += 1

        name = " ".join(dict.fromkeys(p for p in name_parts if p)).strip(" -–—")
        type_str = "; ".join(dict.fromkeys(t for t in type_parts if t)).strip(" -–—")
        name = _trim_footer(name)
        type_str = _trim_footer(type_str)
        if not name and type_str and len(type_str) <= 60:
            name, type_str = type_str, ""
        if not name:
            if warn:
                warn(f"{path.name}/{sheet.name} 第 {i} 行备件缺少名称（数量 {qty:g} {unit or '个'}）")
            name = "(未填写名称)"

        items.append(Part(name=name, qty=qty, unit=unit or "个", type=type_str or None))
        i = j
    return items


# ---- number / text helpers --------------------------------------------

def _to_number(value) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("，", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        pass
    m = re.search(r"-?\d+(?:\.\d+)?", text)
    if m:
        try:
            return float(m.group(0))
        except ValueError:
            return None
    return None


def _to_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


# ---- public API -------------------------------------------------------

def parse_source(path, warn=None) -> list:
    """Parse a source Excel file (all sheets) into a list of Part objects.

    Optional callable warn(msg) receives messages about skipped/unparseable rows.
    """
    path = Path(path)
    parts = []
    for sheet in _open_sheets(path):
        header_row, mapping = _find_header_row(sheet)
        if header_row is not None:
            if _is_packed_format(mapping):
                parts.extend(_parse_packed_sheet(sheet, mapping, header_row, path, warn))
                continue
            parts.extend(_parse_standard(sheet, mapping, header_row, path, warn))
            continue
        parts.extend(_parse_packing_sheet(sheet, path, warn))

    if not parts:
        raise ValueError(f"在 {path.name} 中没有解析到备件数据")
    return parts


def parse_sources(paths, warn=None) -> list:
    """Parse one or more source files, merging all parts in order."""
    all_parts = []
    for p in paths:
        all_parts.extend(parse_source(p, warn=warn))
    return all_parts


def detect_vessel(paths) -> Optional[str]:
    """Try to guess the vessel name from the source files (船名 / 船名/单位 / Vessel Name)."""
    for p in paths:
        p = Path(p)
        try:
            sheets = _open_sheets(p)
        except Exception:
            continue
        for sheet in sheets:
            for row in sheet.iter_rows():
                texts = [str(c.value).strip() for c in row if c.value is not None]
                joined = " ".join(texts)

                m = re.search(r"船名\s*[:：]\s*(\S+)", joined)
                if m:
                    return m.group(1)

                for i, t in enumerate(texts):
                    if "船名/单位" in t or "船名／单位" in t:
                        for other in texts[i + 1:]:
                            if other:
                                return other
                        break

                for t in texts:
                    if "vessel name" not in t.lower():
                        continue
                    m = re.search(r"[Vv]essel\s+[Nn]ame\s*[:：]?\s*(.*)", t)
                    val = (m.group(1) if m else "").strip()
                    if not val:
                        # label cell only; the value sits in a sibling cell on this row
                        for other in texts:
                            if other != t and other.lstrip(":"):
                                candidate = other.lstrip(":")
                                pm = re.search(r"\(([^)]*)\)", candidate)
                                if pm and re.search(r"[\u4e00-\u9fff]", pm.group(1)):
                                    return pm.group(1).strip()
                                clean = candidate.split("(")[0].strip()
                                if clean:
                                    return clean
                        continue
                    pm = re.search(r"\(([^)]*)\)", val)
                    if pm and re.search(r"[\u4e00-\u9fff]", pm.group(1)):
                        return pm.group(1).strip()
                    clean = val.split("(")[0].strip()
                    if clean:
                        return clean
    return None


# ---- bilingual vessel name (中英文船名对照表) ----------------------------

_CJK = re.compile(r"[\u4e00-\u9fff]+")

# traditional -> simplified characters commonly found in ship/port/city names
_T2S = {
    "遠": "远", "偉": "伟", "衛": "卫", "萬": "万", "東": "东", "蘭": "兰",
    "灣": "湾", "馬": "马", "學": "学", "華": "华", "電": "电", "機": "机",
    "艦": "舰", "號": "号", "長": "长", "陽": "阳", "亞": "亚", "歐": "欧",
    "廣": "广", "國": "国", "紐": "纽", "約": "约", "倫": "伦", "漢": "汉",
    "羅": "罗", "維": "维", "爾": "尔", "澤": "泽", "內": "内", "園": "园",
    "豐": "丰", "匯": "汇", "駿": "骏", "鵬": "鹏", "龍": "龙", "麗": "丽",
    "書": "书", "畫": "画", "廠": "厂", "場": "场", "館": "馆", "樓": "楼",
    "標": "标", "樣": "样", "權": "权", "歡": "欢", "橋": "桥", "蘇": "苏",
    "薩": "萨", "賓": "宾", "賽": "赛", "貝": "贝", "貨": "货", "車": "车",
    "輪": "轮", "軸": "轴", "銅": "铜", "鋼": "钢", "鐵": "铁", "鈴": "铃",
    "銀": "银", "鏈": "链", "鐘": "钟", "燈": "灯", "點": "点", "熱": "热",
    "愛": "爱", "樂": "乐", "興": "兴", "舊": "旧", "與": "与", "從": "从",
    "後": "后", "來": "来", "見": "见", "現": "现", "視": "视", "親": "亲",
    "覺": "觉", "觀": "观", "規": "规", "則": "则", "製": "制", "複": "复",
    "烏": "乌", "聖": "圣", "塞": "塞", "納": "纳", "島": "岛", "嶼": "屿",
    "鎮": "镇", "瀋": "沈", "連": "连", "臺": "台", "廈": "厦", "門": "门",
    "煙": "烟", "瓊": "琼", "濱": "滨", "寧": "宁", "溫": "温", "營": "营",
    "蘆": "芦", "錦": "锦", "雲": "云", "黃": "黄", "鹽": "盐", "鏟": "铲",
    "蓮": "莲", "綠": "绿", "極": "极", "紅": "红", "裏": "里", "幾": "几",
    "遜": "逊", "萊": "莱", "雙": "双", "蠍": "蝎", "獅": "狮", "寶": "宝",
    "魚": "鱼", "鵑": "鹃", "楊": "杨", "櫻": "樱", "櫚": "榈", "膠": "胶",
    "楓": "枫", "樺": "桦", "腦": "脑", "樸": "朴", "檸": "柠", "蘋": "苹",
    "藍": "蓝", "無": "无", "欖": "榄", "霧": "雾", "龍": "龙", "菠": "菠",
    "蘿": "萝", "絲": "丝", "蔥": "葱", "薑": "姜", "歸": "归", "參": "参",
    "棗": "枣", "翹": "翘", "葉": "叶", "蟲": "虫", "靈": "灵", "窩": "窝",
    "鮑": "鲍", "蝦": "虾", "裝": "装", "紙": "纸", "織": "织", "滌": "涤",
    "綸": "纶", "襪": "袜", "褲": "裤", "襯": "衬", "夾": "夹", "傘": "伞",
    "櫃": "柜", "鋪": "铺", "蓋": "盖", "單": "单", "簾": "帘", "發": "发",
    "櫥": "橱", "凍": "冻", "飯": "饭", "壺": "壶", "盤": "盘", "鉗": "钳",
    "錘": "锤", "鑿": "凿", "鋸": "锯", "鉋": "刨", "銼": "锉", "鋁": "铝",
    "鉛": "铅", "鋅": "锌", "錫": "锡", "鎳": "镍", "鉻": "铬", "鈦": "钛",
    "鎂": "镁", "鎢": "钨", "鉬": "钼", "釩": "钒", "錳": "锰", "鈷": "钴",
    "鉑": "铂", "氫": "氢", "氬": "氩", "鈉": "钠", "鉀": "钾", "鈣": "钙",
    "鋰": "锂", "鈹": "铍", "釔": "钇", "鋯": "锆", "鈮": "铌", "鍀": "锝",
    "釕": "钌", "銠": "铑", "鈀": "钯", "鎘": "镉", "銦": "铟", "銻": "锑",
    "鍶": "锶", "銣": "铷", "銫": "铯", "鋇": "钡", "鑭": "镧", "鈰": "铈",
    "鐠": "镨", "釹": "钕", "釤": "钐", "銪": "铕", "鋱": "铽", "鏑": "镝",
    "鉺": "铒", "銩": "铥", "鐿": "镱", "鑥": "镥", "鉿": "铪", "鉭": "钽",
    "錸": "铼", "鋨": "锇", "銥": "铱", "鉍": "铋", "釙": "钋", "鈁": "钫",
    "鐳": "镭", "錒": "锕", "釷": "钍", "鏷": "镤", "鈾": "铀", "鎿": "镎",
    "鈽": "钚", "鋂": "镅", "鋦": "锔", "鉳": "锫", "鉲": "锎", "鎄": "锿",
    "鐨": "镄", "鍆": "钔", "鍩": "锘", "鐒": "铹", "鑪": "𬬻", "𨧀": "𬭊",
    "𨭎": "𬭳", "𨨏": "𬭛", "𨭆": "𬭶", "錀": "𬬭", "鎶": "鿔", "鉨": "鿭",
    "鈇": "𫓧", "鏌": "镆", "鉝": "𫟷",
}


def _zh_key(s: str) -> str:
    """Normalize a Chinese name: fullwidth->halfwidth, strip spaces, trad->simplified."""
    s = unicodedata.normalize("NFKC", str(s)).replace(" ", "").replace("\u3000", "")
    return "".join(_T2S.get(ch, ch) for ch in s)


def _en_key(s: str) -> str:
    """Normalize an English/romanized name: uppercase, keep letters/digits only."""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def _split_zh_en(val: str):
    """Split a vessel value like 'COSMERRY LAKE(遠怡湖)' or '远怡湖COSMERRY LAKE' into (chinese, english)."""
    val = str(val).strip().lstrip(":")
    m = re.search(r"[（(]([^）)]*)[)）]", val)
    zh = None
    if m and re.search(r"[\u4e00-\u9fff]", m.group(1)):
        zh = m.group(1).strip()
        val = (val[: m.start()] + val[m.end():]).strip()
    else:
        m = re.search(_CJK, val)
        if m:
            zh = m.group(0).strip()
            val = (val[: m.start()] + val[m.end():]).strip()
    en = val.strip().strip(": ")
    en = re.sub(r"[^A-Za-z0-9\u4e00-\u9fff]+$", "", en)  # strip trailing punctuation
    if zh is None and not en:
        return None, None
    return zh, en or None


def detect_vessel_pair(paths):
    """Return (chinese, english) vessel names found in the sources, or (None, None)."""
    for p in paths:
        p = Path(p)
        try:
            sheets = _open_sheets(p)
        except Exception:
            continue
        for sheet in sheets:
            for row in sheet.iter_rows():
                texts = [str(c.value).strip() for c in row if c.value is not None]
                joined = " ".join(texts)

                m = re.search(r"船名\s*[:：]\s*(\S+)", joined)
                if m:
                    zh, en = _split_zh_en(m.group(1))
                    if zh or en:
                        return zh, en

                for i, t in enumerate(texts):
                    if "船名/单位" in t or "船名／单位" in t:
                        for other in texts[i + 1:]:
                            if other:
                                return _split_zh_en(other)
                        break

                for i, t in enumerate(texts):
                    if "vessel name" not in t.lower():
                        continue
                    m = re.search(r"[Vv]essel\s+[Nn]ame\s*[:：]?\s*(.*)", t)
                    val = (m.group(1) if m else "").strip()
                    if not val:
                        for other in texts[i + 1:]:
                            if not other.lstrip(":"):
                                continue
                            res = _split_zh_en(other.lstrip(":"))
                            if res[0] or res[1]:
                                return res
                        continue
                    res = _split_zh_en(val)
                    if res[0] or res[1]:
                        return res
    return None, None


def load_vessel_names(path):
    """Load the 中英文船名 lookup workbook -> (zh2en, en2zh) name mappings."""
    path = Path(path)
    zh2en, en2zh = {}, {}
    for sheet in _open_sheets(path):
        rows = list(sheet.iter_rows())
        if not rows:
            continue
        header = rows[0]
        pair_cols = []
        for c in header:
            h = _clean(c.value)
            if not h or ("中文" not in h and h != "船名"):
                continue
            for c2 in header:
                if c2.column <= c.column:
                    continue
                h2 = _clean(c2.value)
                if not h2:
                    continue
                if "英文" in h2 or "拼音" in h2:
                    pair_cols.append((c.column, c2.column))
                    break
        for r in rows[1:]:
            for cn_col, en_col in pair_cols:
                cn = _cell_in_row(r, cn_col)
                en = _cell_in_row(r, en_col)
                if cn and en:
                    zh2en.setdefault(_zh_key(cn), _clean(en))
                    en2zh.setdefault(_en_key(en), _zh_key(cn))
    return zh2en, en2zh


def _cell_in_row(row, col):
    for c in row:
        if c.column == col:
            return _clean(c.value)
    return None


def bilingual_vessel(vessel, zh2en, en2zh, english=None, chinese=None):
    """Return a '中文 英文' string when the vessel can be resolved via the lookup table.

    If an English name was read from the source it is matched against the lookup
    first (handles traditional/simplified Chinese differences). Otherwise the
    Chinese name is matched (with a light traditional->simplified conversion).
    Unresolvable names are returned unchanged.
    """
    if english:
        en_key = _en_key(english)
        if en_key in en2zh:
            zh = en2zh[en_key]
            en = zh2en.get(zh, _clean(english))
            return f"{zh} {en}"
        zh = _clean(chinese or vessel or "")
        return f"{zh} {_clean(english)}".strip()
    zh = _clean(chinese or vessel or "")
    key = _zh_key(zh)
    if key in zh2en:
        return f"{key} {zh2en[key]}"
    return zh.strip()