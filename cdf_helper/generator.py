"""Generate the CDF (报关清单) workbook by filling the user's template.

The template layout (Sheet1):
    A1 (merged A1:G1) : title "船名：<vessel>"
    row 2             : headers 序号/备件名称/数量/单位/重量(KG)/单价/RMB/金额RMB
    rows 3.N         : one row per part (序号 and 金额 are formulas)
    last data row + 1 : 合计 total row (yellow fill, SUM formulas)
"""

import copy
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TEMPLATE_SHEET = "Sheet1"
FIRST_DATA_ROW = 3
HEADER_ROW = 2
LAST_HEADER_COL = 7
TEMPLATE_DATA_CAPACITY = 4  # rows 3..6 in the blank template

# --- Template layout contract (validated up-front, see validate_template) ---
EXPECTED_HEADERS = ["序号", "备件名称", "数量", "单位", "重量(KG)", "单价/RMB", "金额RMB"]
TOTAL_MARKER = "合计"
EXPECTED_TOTAL_ROW = FIRST_DATA_ROW + TEMPLATE_DATA_CAPACITY  # 7
# Columns O/P (15/16) in the blank template carry stray "套"/2 leftovers.
LEFTOVER_COLS = (15, 16)


def _copy_style(src_cell, dst_cell):
    dst_cell._style = copy.copy(src_cell._style)


def _style_row_template(ws, row):
    """Capture the cell styles of one template row (data or total) as a list."""
    return [copy.copy(ws.cell(row, col)._style) for col in range(1, LAST_HEADER_COL + 1)]


def _apply_styles(ws, row, styles):
    for col, style in enumerate(styles, start=1):
        ws.cell(row, col)._style = copy.copy(style)


def _validate_template_layout(wb, sheet_name=TEMPLATE_SHEET):
    """Raise ``ValueError`` if the loaded template's layout deviates from what
    :func:`generate` hardcodes (sheet name, header row, total-row position).

    The generator writes data to absolute positions and never re-reads the
    template's header row to discover column mapping.  A mismatched template
    therefore produces a *silently corrupt* workbook rather than an obvious
    failure, so we fail fast with an actionable message instead.
    """
    # --- sheet name ---
    if sheet_name not in wb.sheetnames:
        names = ", ".join(wb.sheetnames) if wb.sheetnames else "（空）"
        raise ValueError(
            f'模板缺少名为"{sheet_name}"的工作表，无法填充。'
            f'当前工作表：{names}。'
            '请使用标准报关清单模板（工作表名为 Sheet1）。'
        )

    ws = wb[sheet_name]

    # --- header row: every column must carry the expected label ---
    mismatches = []
    for col in range(1, LAST_HEADER_COL + 1):
        expected = EXPECTED_HEADERS[col - 1]
        raw = ws.cell(HEADER_ROW, col).value
        actual = str(raw).strip() if raw is not None else ""
        if actual != expected:
            mismatches.append(
                f"{get_column_letter(col)}{HEADER_ROW} 期望“{expected}” 实际“{actual}”"
            )
    if mismatches:
        raise ValueError(
            f'模板表头（第 {HEADER_ROW} 行）与标准格式不符，'
            f'无法确定列映射。{", ".join(mismatches)}。'
            '请使用标准报关清单模板，第 2 行为：'
            + " / ".join(EXPECTED_HEADERS) + '。'
        )

    # --- total-row marker must sit at the expected position ---
    total_cell = ws.cell(EXPECTED_TOTAL_ROW, 1)
    total_val = str(total_cell.value).strip() if total_cell.value is not None else ""
    if total_val != TOTAL_MARKER:
        raise ValueError(
            f'模板第 {EXPECTED_TOTAL_ROW} 行（应为“{TOTAL_MARKER}” 合计行）'
            f'实际为“{total_val}”。生成的合计行会错位，'
            '请使用标准报关清单模板。'
        )


def validate_template(template_path, sheet_name=TEMPLATE_SHEET):
    """Standalone pre-flight check (does **not** mutate the file).

    Returns ``None`` when the template is acceptable, or a human-readable
    error string.  Used by the web layer to give early feedback before
    source files are parsed.
    """
    template_path = Path(template_path)
    if not template_path.is_file():
        return f'模板文件不存在：{template_path}。'

    suffix = template_path.suffix.lower()
    if suffix == ".xls":
        return (
            f'模板文件“{template_path.name}”是旧版 .xls 格式，'
            '仅支持 .xlsx 模板。请转换为 .xlsx 后重试，'
            '或使用服务器提供的标准报关清单模板。'
        )
    if suffix != ".xlsx":
        return (
            f'模板文件“{template_path.name}”格式不受支持（{suffix}）。'
            '请上传 .xlsx 报关清单模板。'
        )

    try:
        wb = load_workbook(template_path, data_only=False, read_only=True)
    except Exception as exc:
        return f'无法读取模板文件：{exc}。'

    try:
        _validate_template_layout(wb, sheet_name)
    except ValueError as exc:
        return str(exc)

    return None


def generate(template_path, parts, vessel_name, output_dir, output_name,
             include_spec=True, sheet_name=TEMPLATE_SHEET):
    """Fill the template with parts and save to output_dir/output_name.

    Returns the absolute path of the generated file.
    """
    template_path = Path(template_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path, data_only=False)
    _validate_template_layout(wb, sheet_name)
    ws = wb[sheet_name]

    # Title cell
    ws.cell(1, 1).value = f"船名：{vessel_name}"

    data_styles = _style_row_template(ws, FIRST_DATA_ROW)
    total_row_in_template = FIRST_DATA_ROW + TEMPLATE_DATA_CAPACITY  # 7
    total_styles = _style_row_template(ws, total_row_in_template)

    # Clear the sample rows 3..6 (their styles are reused via data_styles)
    for r in range(FIRST_DATA_ROW, FIRST_DATA_ROW + TEMPLATE_DATA_CAPACITY):
        for col in range(1, LAST_HEADER_COL + 1):
            ws.cell(r, col).value = None

    # Insert extra rows before the total row if we have more parts than capacity
    n = len(parts)
    if n > TEMPLATE_DATA_CAPACITY:
        ws.insert_rows(total_row_in_template, n - TEMPLATE_DATA_CAPACITY)

    total_row = FIRST_DATA_ROW + n

    for i, part in enumerate(parts):
        r = FIRST_DATA_ROW + i
        ws.cell(r, 1).value = "=ROW()-2"
        ws.cell(r, 2).value = _display_name(part, include_spec)
        ws.cell(r, 3).value = part.qty
        ws.cell(r, 4).value = part.unit
        ws.cell(r, 5).value = part.weight
        ws.cell(r, 6).value = part.price
        ws.cell(r, 7).value = f"=C{r}*F{r}"
        _apply_styles(ws, r, data_styles)
        ws.row_dimensions[r].height = 20

    # Total row
    ws.cell(total_row, 1).value = "合计"
    ws.cell(total_row, 2).value = None
    ws.cell(total_row, 3).value = "=SUM(C$3:INDEX(C:C,ROW()-1))"
    ws.cell(total_row, 4).value = None
    ws.cell(total_row, 5).value = "=SUM(E$3:INDEX(E:E,ROW()-1))"
    ws.cell(total_row, 6).value = None
    ws.cell(total_row, 7).value = "=SUM(G$3:INDEX(G:G,ROW()-1))"
    _apply_styles(ws, total_row, total_styles)
    ws.row_dimensions[total_row].height = 20

    # Tidy up template leftovers outside the table (row 7 cols O/P)
    for col in LEFTOVER_COLS:
        cell = ws.cell(total_row, col)
        if cell.value is not None:
            cell.value = None

    out_path = output_dir / output_name
    wb.save(out_path)
    return str(out_path)


def _display_name(part, include_spec):
    name = part.name.strip()
    if include_spec and part.type:
        spec = part.type.strip()
        if spec and spec != name:
            return f"{name} {spec}"
    return name


def sanitize_filename(name, fallback="file"):
    """Remove characters that are invalid in Windows filenames."""
    cleaned = re.sub(r'[\\/:*?"<>|\r\n\t]', "-", str(name)).strip()
    cleaned = cleaned.strip(" .")
    return cleaned or fallback