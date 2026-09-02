"""Tests for the packing-list (English Receipt/Packing List) parser, incl. a
best-effort check against the real uploaded file when present."""
import sys
import tempfile
import zipfile
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook

from cdf_helper.parser import (
    parse_source,
    parse_sources,
    detect_vessel,
    detect_vessel_pair,
    load_vessel_names,
    bilingual_vessel,
)


def _write_workbook(path: Path):
    wb = Workbook()

    # P1: Sheet1-like, merged header in A, item/qty/name in separate columns + Type line
    ws = wb.active
    ws.title = "P1"
    ws["A1"] = "Item      Quantity(Unit)         Particulars"
    ws["A2"] = "1"
    ws["B2"] = "1  PCE"
    ws["C2"] = "消防泵出海阀(蝶阀)"
    ws["B3"] = "Type:5k250"
    ws["A4"] = "2"
    ws["B4"] = "2 PCE"
    ws["C4"] = "液压法兰式蝶阀"
    ws["B6"] = "** End of Listing **"

    # P2: Sheet3-like, header split across A/B, qty+name in same cell, Type line after
    ws = wb.create_sheet("P2")
    ws["A1"] = "Item"
    ws["B1"] = "Quantity(Unit)          Particulars"
    ws["A2"] = "1"
    ws["B2"] = "3  PC  时间继电器"
    ws["B3"] = "Type:OMRON H3CR-H8L AC220V 50/60HZ"
    ws["A4"] = "2"
    ws["B4"] = "3 PCE"
    ws["C4"] = "时间继电器"
    ws["B5"] = "Type:OMRON H3CR-A8E"
    ws["B7"] = "**End of Listing**"

    # P3: Sheet8-like, qty+name in one merged cell, Type line is the only real name source
    ws = wb.create_sheet("P3")
    ws["A1"] = "Item       Quantity(Unit)          Particulars"
    ws["A2"] = "  1              14   PCE                         plate"
    ws["B2"] = "51506-04H-160"
    ws["D3"] = "Type: 辅机滑油冷却器换热片"
    ws["A4"] = "**   End   of   Listing**"

    # P4: footer text glued onto a name must be trimmed
    ws = wb.create_sheet("P4")
    ws["A1"] = "Item Quantity(Unit) Particulars"
    ws["A2"] = "1"
    ws["B2"] = "4 PCE"
    ws["C2"] = "Sealing Ring 蓋章及簸收 httn://10.13.30.15:8080/x"
    ws["B3"] = "Type:07.02.01.063"
    ws["A5"] = "** End of Listing **"

    # NoHeader: no Item header -> skipped, but carries a Vessel Name
    ws = wb.create_sheet("NoHeader")
    ws["A1"] = "Vessel Name :VESSEL-A(测试船)"
    ws["A2"] = "1  2 PCE  某某"

    wb.save(path)


def test_synthetic():
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "packing.xlsx"
        _write_workbook(path)

        warns = []
        parts = parse_source(path, warn=warns.append)

        # P1 2 items + P2 2 items + P3 1 item + P4 1 item; NoHeader skipped
        assert len(parts) == 6, [p.name for p in parts]
        assert any("未找到 Item/Quantity" in w for w in warns), warns

        p1, p2, p3, p4, p5, p6 = parts
        assert p1.name == "消防泵出海阀(蝶阀)" and p1.qty == 1 and p1.unit == "PCE"
        assert p1.type == "5k250", p1.type
        assert p2.name == "液压法兰式蝶阀" and p2.qty == 2
        assert p3.name == "时间继电器" and p3.qty == 3 and p3.unit == "PC"
        assert p3.type == "OMRON H3CR-H8L AC220V 50/60HZ", p3.type
        assert p4.name == "时间继电器" and p4.qty == 3 and p4.type == "OMRON H3CR-A8E"
        assert p5.qty == 14 and p5.unit == "PCE"
        assert p5.type == "辅机滑油冷却器换热片", p5.type
        assert p6.name == "Sealing Ring", p6.name  # footer trimmed
        assert p6.qty == 4 and p6.type == "07.02.01.063"

        assert detect_vessel([path]) == "测试船"
        assert detect_vessel_pair([path]) == ("测试船", "VESSEL-A")


def _write_embedded_end_workbook(path: Path):
    """A packing list whose '** End of Listing **' text is embedded inside a data
    row's cell (e.g. a Serial No. cell) rather than sitting on its own footer row."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Item"
    ws["B1"] = "Quantity(Unit)"
    ws["C1"] = "Description"
    ws["D1"] = "Serial No"
    ws["A2"] = "1"
    ws["B2"] = "24 PCS"
    ws["C2"] = "CASMSA-26030-0000 Main Contactor Contact"
    ws["D2"] = (
        "Serial No. : 7251785\n"
        "(Winch) head anchor windlass\n"
        "** End of Listing **\n"
        "Signature and Chop"
    )
    wb.save(path)


def test_end_of_listing_embedded_in_data():
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "embedded_end.xlsx"
        _write_embedded_end_workbook(path)

        warns = []
        parts = parse_source(path, warn=warns.append)

        assert len(parts) == 1, parts
        assert parts[0].name == "CASMSA-26030-0000 Main Contactor Contact", parts[0].name
        assert parts[0].qty == 24 and parts[0].unit == "PCS", parts[0]
    print("embedded End of Listing test OK")


def test_vessel_lookup():
    lookup = Path("中英文船名25-5-14.xls")
    if not lookup.exists():
        print("SKIP: 中英文船名 lookup not present")
        return
    zh2en, en2zh = load_vessel_names(lookup)
    assert len(zh2en) == len(en2zh) > 100
    assert zh2en["远怡湖"] == "COSMERRY LAKE"
    assert zh2en["中远安特卫普"] == "COSCO ANTWERP"
    assert en2zh["COSMERRYLAKE"] == "远怡湖"
    assert en2zh["COSCOANTWERP"] == "中远安特卫普"

    # English match handles traditional/simplified differences in the source
    assert bilingual_vessel("中遠安特偉普", zh2en, en2zh, english="coscO ANTWERP", chinese="中遠安特偉普") == "中远安特卫普 COSCO ANTWERP"
    assert bilingual_vessel("遠怡湖", zh2en, en2zh, english="cOSMERRY LAKE", chinese="遠怡湖") == "远怡湖 COSMERRY LAKE"
    # Chinese-only match through the lookup (typed forms)
    assert bilingual_vessel("远怡湖", zh2en, en2zh) == "远怡湖 COSMERRY LAKE"
    assert bilingual_vessel("遠怡湖", zh2en, en2zh) == "远怡湖 COSMERRY LAKE"
    # unresolvable -> unchanged
    assert bilingual_vessel("测试船", zh2en, en2zh) == "测试船"


def test_real_file_if_present():
    real = Path("uploads/中远安特伟普-大连26-8-20.xlsx")
    if not real.exists():
        print("SKIP: 中远安特伟普 file not present")
        return
    parts = parse_sources([real])
    assert len(parts) > 0
    assert all(p.name for p in parts), "every parsed part must have a name"
    assert any(p.name == "(未填写名称)" for p in parts), "Sheet1 unnamed items expected"
    vessel = detect_vessel([real])
    print(f"real file: {len(parts)} items, vessel={vessel}")

    # second real sample: 9 mixed-format sheets (Chinese 签收单 + Yuantong packing lists)
    real2 = Path("uploads/远怡湖-湛江26-8-20.xlsx")
    if not real2.exists():
        print("SKIP: 远怡湖 file not present")
        return
    parts2 = parse_sources([real2])
    assert len(parts2) == 29, [p.name for p in parts2]
    assert detect_vessel([real2]) == "遠怡湖"
    print(f"real file2: {len(parts2)} items, vessel=遠怡湖")



# ---- packed single-column format tests ----

def _write_packed_workbook(path: Path):
    """Create a workbook with packed single-column format sheets."""
    wb = Workbook()

    # Sheet1: multi-row packed (header on row 5, data on rows 6-7)
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=4, column=1, value="单位： 寰宇船务企业有限公司     远怡湖     单号 SZY268003")
    ws.cell(row=5, column=1, value="序号           设备                       名称                           编号             单位   数量     备注")
    ws.cell(row=6, column=1, value="1        左舷梯             STOWING ROPE          6×37-13.0-170-I           1")
    ws.cell(row=7, column=1, value="2                           HOISHING ROPE          6×37-15.0-180-I           2")
    ws.cell(row=8, column=1, value="目的地：湛江")
    ws.cell(row=9, column=1, value="经办人签字：王叶")

    # Sheet2: multi-line cell format (header + data in same cell)
    ws = wb.create_sheet("Sheet2")
    ws.cell(row=4, column=1,
            value="单位： 寰宇船务企业有限公司 远怡湖 单号 SZY267887\n"
                  "序号           设备                       名称                           编号             单位   数量     备注\n"
                  "  1      防海生物装置     Anode MGC650R 650x110     3/Drg. JCA-11529           2")

    # Sheet3: Chinese equipment + English name in packed format
    ws = wb.create_sheet("Sheet3")
    ws.cell(row=4, column=1, value="序号  设备  名称  编号  单位  数量  备注")
    ws.cell(row=5, column=1, value="1   泵箱总成   液压泵   0420848-92001 / 3   1")
    ws.cell(row=6, column=1, value="2   法兰   管件   Code : 427HL110-0 B   3")
    ws.cell(row=7, column=1, value="日期：2026/8/25")

    wb.save(path)


def test_packed_synthetic():
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "packed.xlsx"
        _write_packed_workbook(path)

        warns = []
        parts = parse_source(path, warn=warns.append)

        # Sheet1: 2 parts; Sheet2: 1 part; Sheet3: 2 parts = 5 total
        assert len(parts) == 5, [(p.name, p.qty, p.type) for p in parts]

        # Sheet1 Row 6: equipment=左舷梯, name=STOWING ROPE, code=6×37, qty=1
        assert parts[0].name == "左舷梯 STOWING ROPE", parts[0].name
        assert parts[0].qty == 1.0
        assert parts[0].type == "6×37-13.0-170-I", parts[0].type

        # Sheet1 Row 7: no equipment, name=HOISHING ROPE, code=6×37, qty=2
        assert parts[1].name == "HOISHING ROPE", parts[1].name
        assert parts[1].qty == 2.0
        assert parts[1].type == "6×37-15.0-180-I", parts[1].type

        # Sheet2 (multi-line cell): equipment=防海生物装置, name=Anode MGC650R 650x110, code=3/Drg. JCA-11529, qty=2
        assert parts[2].name == "防海生物装置 Anode MGC650R 650x110", parts[2].name
        assert parts[2].qty == 2.0
        assert parts[2].type == "3/Drg. JCA-11529", parts[2].type

        # Sheet3: qty should be the last numeric value
        assert parts[3].qty == 1.0
        assert parts[3].type == "0420848-92001 / 3", parts[3].type
        assert parts[4].qty == 3.0

        # Footer lines should not become parts
        assert not any("目的地" in p.name for p in parts), "footer leaked into parts"
        assert not any("签字" in p.name for p in parts), "footer leaked into parts"
        assert not any("日期" in p.name for p in parts), "footer leaked into parts"
        print("packed format synthetic test OK")


def test_packed_real_file_if_present():
    real = Path("远怡湖missing sheets.xlsx")
    if not real.exists():
        print("SKIP: 远怡湖missing sheets.xlsx not present")
        return
    parts = parse_source(real, warn=lambda m: print(f"  WARN: {m}"))
    assert len(parts) > 0, "expected parts from packed format file"
    assert all(p.name for p in parts), "every parsed part must have a name"
    print(f"packed real file: {len(parts)} items")
    for p in parts:
        print(f"  name={p.name!r}, qty={p.qty}, type={p.type!r}")

def _write_wrapped_header_workbook(path: Path):
    """Sheet whose qty header label is wrapped across newlines, mirroring
    远新湖-加单.xlsx Sheet6 ('订单数\\n量\\n(NUM)')."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet6"
    ws["A1"] = "备件送船签收单"
    ws["A2"] = "远新湖 SHIP"
    ws["A3"] = "序号"
    ws["B3"] = "订单数\n量\n(NUM)"
    ws["C3"] = "单位(UNIT)"
    ws["D3"] = "备件编号(PARTS.NO)"
    ws["E3"] = "备件名称(PARTS.NAME)"
    ws["G3"] = "型号与规格(SPECIFICATION)"
    ws["A4"] = "1"; ws["B4"] = "1"; ws["C4"] = "PC"; ws["E4"] = "Circuit Breaker"
    ws["A5"] = "2"; ws["B5"] = "2"; ws["C5"] = "PC"; ws["E5"] = "Contactor"
    wb.save(path)


def test_multiline_wrapped_header():
    """A 数量 column label wrapped across lines (e.g. '订\n订单数\n量\n(NUM)')
    must still be recognized as the qty header. Regression for 远新湖-加单.xlsx
    Sheet6 being skipped entirely."""
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "wrapped.xlsx"
        _write_wrapped_header_workbook(path)

        warns = []
        parts = parse_source(path, warn=warns.append)

        assert len(parts) == 2, [(p.name, p.qty) for p in parts]
        assert parts[0].name == "Circuit Breaker" and parts[0].qty == 1 and parts[0].unit == "PC"
        assert parts[1].name == "Contactor" and parts[1].qty == 2 and parts[1].unit == "PC"
        assert not any("跳过" in w for w in warns), warns
    print("multiline wrapped header test OK")


def _write_xlsx_with_bad_number_cell(path: Path):
    """A valid parts workbook, then inject a number-typed cell holding a stray
    '.' -- the exact defect that makes openpyxl 3.1 raise
    'ValueError: could not convert string to float' and abort the whole load."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["序号", "备件名称", "数量", "单位"])
    ws.append([1, "Circuit Breaker", 1, "PC"])
    ws.append([2, "Contactor", 2, "PC"])
    tmp = path.with_name(path.name + ".tmp")
    wb.save(tmp)
    with zipfile.ZipFile(tmp) as zin, zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item)
            if item.filename.endswith("worksheets/sheet1.xml"):
                xml = data.decode("utf-8")
                xml = xml.replace(
                    "</sheetData>",
                    '<row r="9"><c r="A9"><v>.</v></c></row></sheetData>',
                    1,
                )
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    tmp.unlink()


def test_bad_number_cell_xlsx_loads():
    """openpyxl 3.1 raises 'could not convert string to float: .' on a stray
    '.' in a number cell, aborting the workbook. _open_sheets must still load
    it (lenient) and parse the parts. Regression for 新茂洋-六横岛26-8-27.xlsx."""
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "badnum.xlsx"
        _write_xlsx_with_bad_number_cell(path)

        warns = []
        parts = parse_source(path, warn=warns.append)

        assert len(parts) == 2, [(p.name, p.qty) for p in parts]
        names = [p.name for p in parts]
        assert "Circuit Breaker" in names and "Contactor" in names, names
        assert not any("跳过" in w for w in warns), warns
    print("bad number cell xlsx test OK")


def _write_deshanghai_workbook(path: Path):
    """德胜海-style Yuantong receipt: two sheets with NO 'Item' column.

    Sheet1 carries the 'Quantity(Unit) Particulars Part No' label row (no Item);
    Sheet2 has that label row missing entirely (header-less). Both carry an
    '**End of Listing**' footer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Quantity(Unit)      Particulars      Part No"
    ws["A2"] = "2 PCE"
    ws["B2"] = "压侧传感器(Pressure"
    ws["C2"] = "Tranmitter)"
    ws["D2"] = "02.01.02.017"
    ws["A3"] = "Type:8298.28.2517"
    ws["B3"] = "Rang:0-10bar"
    ws["A5"] = "** End of Listing **"
    ws = wb.create_sheet("Sheet2")
    ws["A1"] = "DE SHENG HAI RT26026345  Page 1 of 1"
    ws["A3"] = "Part No"
    ws["A5"] = "2 PCE"
    ws["B5"] = "光强传感器"
    ws["C5"] = "02.01.01.017"
    ws["A7"] = "** End of Listing **"
    wb.save(path)


def test_deshanghai_no_item_header_and_headerless():
    """Regression for 德胜海.xlsx: sheets with no 'Item' column, one with a
    'Quantity(Unit) Particulars Part No' header and one header-less, both
    bounded by '**End of Listing**'."""
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "deshanghai.xlsx"
        _write_deshanghai_workbook(path)
        warns = []
        parts = parse_source(path, warn=warns.append)
        assert len(parts) == 2, [(p.name, p.qty) for p in parts]

        header_part, headerless_part = parts
        # Sheet1 (header present, no Item): qty parsed, Part No code dropped
        assert header_part.qty == 2 and header_part.unit == "PCE"
        assert "压侧传感器" in header_part.name
        assert header_part.type == "8298.28.2517"
        assert "02.01.02.017" not in header_part.name
        # Sheet2 (header-less): recovered from first qty row up to End of Listing
        assert headerless_part.qty == 2 and headerless_part.unit == "PCE"
        assert "光强传感器" in headerless_part.name
        assert "02.01.01.017" not in headerless_part.name
    print("deshanghai no-item + headerless test OK")


def _write_yuantong_ocr_workbook(path: Path):
    """远棠湾-加单.xlsx Sheet6 layout: Yuantong receipt whose qty header is OCR-corrupted
    to 'Qantily' (dist 2 from 'Quantity') and whose footer is merged as
    'End ofListing' (no space). Columns: A=seq, B=Qty(Unit), D=Item, G=Part No."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet6"
    ws["A1"] = "技术服务中心"
    ws["A14"] = 28.0
    ws["B14"] = "Qantily"
    ws["D14"] = "Item"
    ws["G14"] = "Part No"
    rows = [
        (29.0, "1 PC",  "Pin",                          57517501.0),
        (30.0, "30 PC", "Hexagon Head Serew",           57617301.0),
        (31.0, "30 PC", "Lock Washer",                  57627901.0),
        (32.0, "5 PC",  "Guide Ring",                   59777703.0),
        (33.0, "2 PC",  "Gaskel\n3-FAY VALVE SPARE PARIS DN25", 59801801.0),
        (34.0, "2 PC",  "INCLIDING 59801801-08\nBody Gasket",
                                                       "906691701 had substituted\n  901639601"),
    ]
    for r, (seq, qty, name, pno) in enumerate(rows, start=15):
        ws.cell(row=r, column=1, value=seq)
        ws.cell(row=r, column=2, value=qty)
        ws.cell(row=r, column=4, value=name)
        ws.cell(row=r, column=7, value=pno)
    ws["E21"] = '"End ofListing*'
    ws["F22"] = "Signature and Chop"
    wb.save(path)


def test_yuantong_ocr_header_and_merged_footer():
    """Regression for 远棠湾-加单.xlsx: OCR 'Qantily' header + merged
    'End ofListing' footer must still be recognised and parsed."""
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "yuantong_ocr.xlsx"
        _write_yuantong_ocr_workbook(path)
        warns = []
        parts = parse_source(path, warn=warns.append)

        assert len(parts) == 6, [(p.name, p.qty) for p in parts]
        by = {p.name: p for p in parts}
        assert "Pin" in by and by["Pin"].qty == 1 and by["Pin"].unit == "PC"
        assert by["Hexagon Head Serew"].qty == 30 and by["Hexagon Head Serew"].unit == "PC"
        assert by["Lock Washer"].qty == 30
        assert by["Guide Ring"].qty == 5
        assert any("3-FAY VALVE" in n and p.qty == 2 for n, p in by.items())
        # Part-No column (G) is excluded -> codes must not leak into names
        assert all("57517501" not in p.name and "57617301" not in p.name for p in parts), parts
        assert not any("跳过" in w for w in warns), warns
    print("yuantong OCR header + merged footer test OK")


def _write_yuanshou_workbook(path: Path):
    """远秋湖-加单.xlsx format: a Chinese ship-material delivery receipt whose
    qty column is labelled `发货量` (not `数量`), with `船舶物资名称` (name) and
    `单价` (price) headers. Footer rows like `本页合计`/`收货人签字` must be excluded."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["行号", "船舶物资名称", "船舶物资规格/品牌", "单位", "发货量", "单价", "备注"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    data = [
        ("#1", "打印机硒鼓", "BROTHER MFC-7860DN", "个", 4, 1058.33, ""),
        ("#2", "HDIM高清线", "通用", "根", 2, 46, ""),
        ("#3", "打印机硒鼓", "HP LaserJet M128fn", "只", 20, 735, ""),
    ]
    for r, row in enumerate(data, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    # footer rows (no part name -> must be excluded)
    ws.cell(row=5, column=1, value="本页合计")
    ws.cell(row=5, column=6, value=22232.14)
    ws.cell(row=6, column=1, value="收货人签字")
    ws.cell(row=6, column=7, value="盖章")
    wb.save(path)


def test_yuanshou_delivery_receipt_format():
    """Regression: `发货量` qty column + `船舶物资名称`/`单价` headers on a
    ship-material delivery receipt (远秋湖-加单.xlsx). Before the fix, `发货量`
    wasn't recognized as the qty column -> no header -> parse_source raised
    ValueError('...没有解析到备件数据')."""
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "yuanshou.xlsx"
        _write_yuanshou_workbook(path)
        warns = []
        parts = parse_source(path, warn=warns.append)

        assert len(parts) == 3, [(p.name, p.qty, p.price) for p in parts]
        names = [(p.name, p.qty, p.unit, p.price) for p in parts]
        assert ("打印机硒鼓", 4.0, "个", 1058.33) in names, names
        assert ("HDIM高清线", 2.0, "根", 46.0) in names, names
        assert ("打印机硒鼓", 20.0, "只", 735.0) in names, names
        # footers excluded (no name -> never emitted as a part)
        assert all("合计" not in (p.name or "") and p.name != "收货人签字" for p in parts), parts
        # no header-skip / row-damage warnings: the sheet WAS recognised as a
        # Chinese parts header and every data row parsed cleanly.
        assert not any("未找到" in w or "跳过" in w or "缺少" in w for w in warns), warns
    print("yuanshou delivery-receipt format test OK")


if __name__ == "__main__":
    test_synthetic()
    test_end_of_listing_embedded_in_data()
    test_deshanghai_no_item_header_and_headerless()
    test_yuantong_ocr_header_and_merged_footer()
    test_yuanshou_delivery_receipt_format()
    test_packed_synthetic()
    test_multiline_wrapped_header()
    test_bad_number_cell_xlsx_loads()
    test_vessel_lookup()
    test_packed_real_file_if_present()
    test_real_file_if_present()
    print("\nALL PARSER TESTS PASSED")